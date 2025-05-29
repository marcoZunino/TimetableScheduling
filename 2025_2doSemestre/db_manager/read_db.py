from .db_config import *
from openpyxl import Workbook


# leer materias
def read_materias(connection):
    cursor = connection.cursor()
    # cursor.execute("SELECT id, nombre FROM materias")
    # rows = cursor.fetchall()
    # materias = []
    # for row in rows:
    #     materias.append({"id": row[0], "nombre": row[1]})
    # return materias

# leer grupos

# leer profesores
def read_profesores(connection):
    cursor = connection.cursor()
    cursor.execute("SELECT nombre, nombre_completo, min_max_dias FROM profesores")
    rows = cursor.fetchall()
    profesores = []
    for row in rows:
        nombre = row[0]
        cursor.execute(f"""
                    SELECT b.dia, b.hora_inicio, p.valor
                    FROM prioridades p, profesores prof, bloques_horarios b
                    WHERE p.profesor = prof.cedula
                    AND prof.nombre = '{nombre}'
                    AND p.bloque_horario = b.id
                    """)
        prioridades = cursor.fetchall()
        profesores.append([r for r in row] + [prioridades])
    return profesores

def read_prioridades(connection, profesor):
    #  -> list[dia.nombre, hora_inicio, valor]
    
    cursor = connection.cursor()
    cursor.execute(f"""
                SELECT b.dia, b.hora_inicio, p.valor
                FROM prioridades p, bloques_horarios b
                WHERE p.profesor = '{profesor}'
                AND p.bloque_horario = b.id
                """)
    prioridades = cursor.fetchall()
    cursor.close()
    return prioridades

def read_last_update(connection, profesor):
    cursor = connection.cursor()
    cursor.execute(f"""
                SELECT ultima_modificacion::timestamp 
                   AT TIME ZONE 'UTC' 
                   AT TIME ZONE 'America/Montevideo',
                   min_max_dias
                FROM profesores
                WHERE nombre = '{profesor}'
                """)
    row = cursor.fetchone()
    cursor.close()

    last_update = row[0]
    min_max_dias = "min" if row[1] else None

    return last_update, min_max_dias



# leer horarios
def read_horarios(connection: psycopg2.extensions.connection):
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM horarios")

    rows = cursor.fetchall()

    horarios = []

    for inicio, fin in rows:
        cursor.execute(f"""
                    SELECT t.turno, t.excepcional
                    FROM turnos_horarios t
                    WHERE t.hora_inicio = '{inicio}'
                    AND t.hora_fin = '{fin}'
                    """)
        data = cursor.fetchall()
        turnos = [t[0] for t in data]
        turnos_excepcional = [t[0] for t in data if t[1]]

        horarios.append((inicio, fin, turnos, turnos_excepcional))

    cursor.close()

    return horarios

def read_turnos(connection):
    cursor = connection.cursor()
    cursor.execute("SELECT nombre FROM turnos")
    rows = cursor.fetchall()
    cursor.close()
    return [row[0] for row in rows]



def write_prioridades_excel(dias, horarios, profesores):
    
    # Crear un libro de trabajo y una hoja
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Prioridades"

    # Escribir encabezados
    sheet["A1"] = "Profesor"
    for i, dia in enumerate(dias):
        sheet.cell(row=1, column=i + 2, value=dia)

    # # Escribir datos de profesores y prioridades
    # for i, profesor in enumerate(profesores):
    #     sheet.cell(row=i + 2, column=1, value=profesor.nombre)
    #     for j, dia in enumerate(dias):
    #         if dia in profesor.prioridades:
    #             sheet.cell(row=i + 2, column=j + 2, value="X")

    # Guardar el archivo
    workbook.save("prioridades.xlsx")



